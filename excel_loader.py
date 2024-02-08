from openpyxl import load_workbook

# def conflict_resolution(values, index):
#     region_values = values.items()
#     for i in enumerate(region_values, start=1):
#         if region_values[i][1] == region_values[i-1][1] and region_values[i][0] < region_values[i-1][0]:
#             pass

def get_param_name(row_index, sheet):
    param_names = []
    for param in sheet[row_index]:
        if param.value == None:
            break
        param_names.append(param.value)

    return param_names

def get_sheet_data(sheet):
    param_names = get_param_name(3, sheet)

    sred = {}
    i = 4

    while sheet[i][0].value != None:

        count = 0
        summ = 0

        for column in range(2, len(param_names), 2):
            if sheet[i][column+1].value == None:
                continue

            count += 1
            summ += sheet[i][column+1].value

        sred[sheet[i][0].value] = round(summ/count,2)
        i += 1

    rank_values = dict(sorted(sred.items(), key=lambda item: item[1], reverse=True))
    rank_index = {raion:index for index, (raion, value) in enumerate(rank_values.items()) }

    leader_name = list(rank_values.items())[0][0]
    leader_value = list(rank_values.items())[0][1]

    row_dat = []
    i = 4
    while sheet[i][0].value != None:
        for column in range(2, len(param_names), 2):
            if sheet[i][column + 1].value == None:
                continue

            row = []

            row.append(sheet[i][0].value)  # Наименование МО
            row.append(sheet.title)  # Направление
            row.append(param_names[column])  # Показатель
            row.append(sheet[i][column].value)  # Значение
            row.append(sheet[i][column + 1].value)  # Балл
            row.append(leader_name)  # Лидер
            row.append(leader_value)  # Значение лидера
            row.append(rank_index[sheet[i][0].value])  # Место

            row.append(sheet[i][1].value)  # Период

            row_dat.append(row)

        i += 1

    return {"raw_data": row_dat, 'rank_values':rank_values, 'rank_index':rank_index }

def get_src_data(filename):
    xls_datd = load_workbook(filename)

    for sheet in xls_datd.worksheets:
        data = get_sheet_data(sheet)
        return data

def get_src_data_index(filename, index=0):
    xls_datd = load_workbook(filename)
    data = get_sheet_data(xls_datd.worksheets[index])
    return data

